# -*- coding: utf-8 -*-
import base64
import io
import logging

from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

# Tempat cetak berdasarkan company_id (sesuai ketentuan PDE KBA)
TEMPAT_BY_COMPANY = {
    1: 'Surabaya',
    2: 'Malang',
    9: 'Surabaya',
    10: 'Trangkil',
}

REQUIRED_COLUMNS = ['NGP', 'REGISTER', 'CAIR', 'GULA_JUAL']

# Jumlah NGP per proses render wkhtmltopdf (1 NGP = 1 halaman A5).
# Batch besar dipecah agar tiap render ringan, lalu hasil PDF digabung.
# Bila render masih berat di server, turunkan angka ini.
CETAK_CHUNK_SIZE = 100

SATUAN = ['', 'SATU', 'DUA', 'TIGA', 'EMPAT', 'LIMA', 'ENAM', 'TUJUH', 'DELAPAN', 'SEMBILAN',
          'SEPULUH', 'SEBELAS']


def angka_ke_terbilang(n):
    """Konversi bilangan bulat non-negatif ke terbilang Bahasa Indonesia (huruf besar)."""
    n = int(round(n or 0))
    if n == 0:
        return 'NOL'

    def convert(num):
        if num < 12:
            return SATUAN[num]
        if num < 20:
            return f"{SATUAN[num - 10]} BELAS"
        if num < 100:
            sisa = num % 10
            return f"{convert(num // 10)} PULUH" + (f" {convert(sisa)}" if sisa else '')
        if num < 200:
            sisa = num % 100
            return "SERATUS" + (f" {convert(sisa)}" if sisa else '')
        if num < 1000:
            sisa = num % 100
            return f"{convert(num // 100)} RATUS" + (f" {convert(sisa)}" if sisa else '')
        if num < 2000:
            sisa = num % 1000
            return "SERIBU" + (f" {convert(sisa)}" if sisa else '')
        if num < 1000000:
            sisa = num % 1000
            return f"{convert(num // 1000)} RIBU" + (f" {convert(sisa)}" if sisa else '')
        if num < 1000000000:
            sisa = num % 1000000
            return f"{convert(num // 1000000)} JUTA" + (f" {convert(sisa)}" if sisa else '')
        sisa = num % 1000000000
        return f"{convert(num // 1000000000)} MILYAR" + (f" {convert(sisa)}" if sisa else '')

    return convert(n).strip()


class KaCetakNgpBatch(models.Model):
    _name = 'ka.cetak.ngp.batch'
    _description = 'Batch Cetak NGP (Nota Gula Petani)'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = 'id desc'

    name = fields.Char(
        string='Nama Batch', required=True, tracking=True,
        default=lambda self: _('Batch NGP Baru'))
    company_id = fields.Many2one(
        'res.company', string='Unit/Company', required=True,
        default=lambda self: self.env.company, tracking=True)

    periode = fields.Char(
        string='Periode', required=True, tracking=True,
        help='Contoh: II B (24 JUNI 2026 S.D 30 JUNI 2026)')
    tanggal_cetak = fields.Date(
        string='Tanggal Cetak', required=True,
        default=fields.Date.context_today, tracking=True)
    tanggal_batas_pengambilan = fields.Date(
        string='Batas Pengambilan', required=True, tracking=True)
    tahun_qr = fields.Char(
        string='Kode Tahun (QR)', compute='_compute_tahun_qr', store=True,
        help='2 digit tahun, otomatis dari Tanggal Cetak. Dipakai untuk Nomor DO & QR Code.')

    tempat = fields.Char(string='Tempat', tracking=True)
    pejabat_nama = fields.Char(string='Nama Penandatangan', tracking=True)
    pejabat_jabatan = fields.Char(
        string='Jabatan', default='Kepala Bagian TUK', tracking=True)

    file = fields.Binary(string='File Excel (.xlsx)', attachment=False)
    filename = fields.Char(string='Nama File')

    line_ids = fields.One2many('ka.cetak.ngp.line', 'batch_id', string='Daftar NGP')
    line_count = fields.Integer(compute='_compute_line_stats', string='Jumlah Baris')
    matched_count = fields.Integer(compute='_compute_line_stats', string='Register Valid')
    unmatched_count = fields.Integer(compute='_compute_line_stats', string='Register Tidak Ditemukan')

    state = fields.Selection([
        ('draft', 'Draft'),
        ('imported', 'Sudah Diimport'),
        ('validated', 'Sudah Divalidasi'),
    ], string='Status', default='draft', tracking=True)

    @api.depends('tanggal_cetak')
    def _compute_tahun_qr(self):
        for rec in self:
            rec.tahun_qr = str(rec.tanggal_cetak.year)[-2:] if rec.tanggal_cetak else ''

    @api.depends('line_ids.is_valid')
    def _compute_line_stats(self):
        for rec in self:
            rec.line_count = len(rec.line_ids)
            rec.matched_count = len(rec.line_ids.filtered('is_valid'))
            rec.unmatched_count = rec.line_count - rec.matched_count

    @api.onchange('company_id')
    def _onchange_company_id(self):
        if self.company_id:
            self.tempat = TEMPAT_BY_COMPANY.get(self.company_id.id, self.company_id.name)

    # ───────────────────────── Import Excel ─────────────────────────
    def action_import_excel(self):
        self.ensure_one()
        try:
            import openpyxl
        except ImportError:
            raise UserError(_(
                "Library 'openpyxl' belum terpasang di server Odoo.\n"
                "Pasang dengan: pip install openpyxl"))

        if not self.file:
            raise UserError(_("Unggah file Excel terlebih dahulu."))

        try:
            data = base64.b64decode(self.file)
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        except Exception as e:
            raise UserError(_("Gagal membuka file Excel: %s") % e)

        ws = wb.worksheets[0]
        header_row = [self._cs(c.value).upper() for c in next(ws.iter_rows(min_row=1, max_row=1))]

        missing = [h for h in REQUIRED_COLUMNS if h not in header_row]
        if missing:
            raise UserError(_(
                "Kolom berikut tidak ditemukan di sheet '%(sheet)s': %(cols)s") % {
                'sheet': ws.title, 'cols': ', '.join(missing)})

        # Kolom 'NGP' muncul dua kali (5-digit & 7-digit) -> pakai index pertama = NGP 5-digit
        col = {}
        for idx, h in enumerate(header_row):
            if h not in col:
                col[h] = idx

        vals_list = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None or all(v in (None, '') for v in row):
                continue
            ngp_nomor = self._cs(row[col['NGP']])
            if not ngp_nomor:
                # NGP kosong -> tidak perlu dicetak, skip baris ini
                continue
            register_kode = self._cs(row[col['REGISTER']])
            vals_list.append({
                'batch_id': self.id,
                'ngp_nomor': ngp_nomor,
                'register_kode': register_kode,
                'gula_jual': self._cf(row[col['GULA_JUAL']]),
                'cair': self._cs(row[col['CAIR']]),
            })

        if not vals_list:
            raise UserError(_("Tidak ada baris data yang valid ditemukan di file Excel."))

        self.line_ids.unlink()
        self.env['ka.cetak.ngp.line'].create(vals_list)
        self.state = 'imported'
        return self._reload()

    # ───────────────────────── Validasi Register ─────────────────────────
    def action_validasi_register(self):
        self.ensure_one()
        if not self.line_ids:
            raise UserError(_("Belum ada data untuk divalidasi. Import Excel terlebih dahulu."))

        Register = self.env['ka.sita.register']
        codes = list(set(self.line_ids.mapped('register_kode')))
        registers = Register.search([('kode_register', 'in', codes)])
        reg_by_code = {r.kode_register: r for r in registers}

        unmatched_codes = []
        for line in self.line_ids:
            reg = reg_by_code.get(line.register_kode)
            line.register_id = reg.id if reg else False
            if not reg:
                unmatched_codes.append(line.register_kode)

        self.state = 'validated'

        if unmatched_codes:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _("Register Tidak Ditemukan"),
                    'message': _("Register %(codes)s belum ada di master register.") % {
                        'codes': ', '.join(sorted(set(unmatched_codes)))},
                    'type': 'warning',
                    'sticky': True,
                    'next': self._reload(),
                },
            }

        return self._reload()

    def action_print_all_valid(self):
        self.ensure_one()
        lines = self.line_ids.filtered('is_valid')
        if not lines:
            raise UserError(_("Tidak ada baris dengan register valid untuk dicetak."))
        return lines._cetak_ngp_pdf(nama_file='NGP_%s' % (self.name or self.id))

    def _reload(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'current',
        }

    @staticmethod
    def _cs(v):
        return str(v).strip() if v not in (None, '') else ''

    @staticmethod
    def _cf(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0


class KaCetakNgpLine(models.Model):
    _name = 'ka.cetak.ngp.line'
    _description = 'Baris NGP (Nota Gula Petani)'
    _order = 'ngp_nomor'

    batch_id = fields.Many2one(
        'ka.cetak.ngp.batch', string='Batch', required=True, ondelete='cascade')
    company_id = fields.Many2one(
        related='batch_id.company_id', store=True, string='Unit/Company')

    ngp_nomor = fields.Char(string='No. NGP', required=True, readonly=True)
    register_kode = fields.Char(string='Kode Register', required=True, readonly=True)
    gula_jual = fields.Float(string='Jumlah Gula (Kg)')
    cair = fields.Char(string='Kode Cair')

    register_id = fields.Many2one(
        'ka.sita.register', string='Register (ka_sita)', ondelete='set null', readonly=True)
    is_valid = fields.Boolean(
        string='Register Valid', compute='_compute_is_valid', store=True)

    # Nama & Desa diambil dari ka.sita.register saat Validasi Register.
    # Kecamatan diambil dari ka.kud (digit pertama Kode Register = Kode KUD).
    # Boleh dikoreksi manual bila perlu; tidak akan ketimpa lagi selama field
    # sumbernya (register_id / register_kode) tidak berubah.
    nama = fields.Char(string='Nama', compute='_compute_dari_register', store=True, readonly=False)
    desa = fields.Char(string='Desa', compute='_compute_dari_register', store=True, readonly=False)
    kecamatan = fields.Char(string='Kecamatan', compute='_compute_kecamatan', store=True, readonly=False)

    terbilang = fields.Char(string='Terbilang', compute='_compute_terbilang', store=True, readonly=False)

    nomor_do = fields.Char(string='Nomor DO', compute='_compute_kode', store=True)
    kode_qr = fields.Char(string='Kode QR', compute='_compute_kode', store=True)
    qr_svg = fields.Html(string='QR SVG', compute='_compute_qr_svg', sanitize=False)

    @api.depends('register_id')
    def _compute_is_valid(self):
        for rec in self:
            rec.is_valid = bool(rec.register_id)

    @api.depends('register_id')
    def _compute_dari_register(self):
        for rec in self:
            reg = rec.register_id
            if reg:
                rec.nama = reg.nama_register or '-'
                rec.desa = reg.desa_id.nama if reg.desa_id else '-'
            # kalau register tidak valid, biarkan nilai sebelumnya (kosong) apa adanya

    @api.depends('register_kode')
    def _compute_kecamatan(self):
        Kud = self.env['ka.kud']
        for rec in self:
            kode = (rec.register_kode or '')[:1]
            kud = Kud.search([('kode', '=', kode)], limit=1) if kode else Kud.browse()
            rec.kecamatan = kud.nama if kud else '-'

    @api.depends('gula_jual')
    def _compute_terbilang(self):
        for rec in self:
            rec.terbilang = f"{angka_ke_terbilang(rec.gula_jual)} KILOGRAM"

    @api.depends('ngp_nomor', 'cair', 'batch_id.tahun_qr')
    def _compute_kode(self):
        for rec in self:
            ngp5 = (rec.ngp_nomor or '').zfill(5)
            cair = rec.cair or ''
            rec.nomor_do = f"DO {ngp5} / {cair}"
            tahun = rec.batch_id.tahun_qr or ''
            rec.kode_qr = f"{tahun}{ngp5}/{cair}" if tahun else ''

    @api.depends('kode_qr')
    def _compute_qr_svg(self):
        for rec in self:
            rec.qr_svg = self._generate_qr_svg(rec.kode_qr) if rec.kode_qr else False

    @staticmethod
    def _generate_qr_svg(data):
        """QR sebagai inline SVG (bukan <img>), supaya wkhtmltopdf tidak membuka
        file descriptor per gambar. Aman untuk cetak batch 1000+ NGP."""
        import qrcode
        import qrcode.image.svg
        qr = qrcode.QRCode(
            version=None,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=10, border=1)
        qr.add_data(data)
        qr.make(fit=True)
        img = qr.make_image(image_factory=qrcode.image.svg.SvgPathImage)
        buf = io.BytesIO()
        img.save(buf)
        svg = buf.getvalue().decode()
        # buang deklarasi XML supaya bisa ditanam inline di HTML
        if svg.startswith('<?xml'):
            svg = svg.split('?>', 1)[1].strip()
        return svg

    # ───────────────────────── Cetak (chunked + progress) ─────────────────────────
    def _cetak_ngp_pdf(self, nama_file='NGP'):
        """Cetak PDF NGP.

        - ≤ CETAK_CHUNK_SIZE baris: render langsung (cepat, tanpa progress).
        - > CETAK_CHUNK_SIZE baris: buka layar progress (client action) yang
          me-render per potongan lewat RPC terpisah lalu menggabungkannya.
          Ini menampilkan animasi loading + progress bar sampai selesai, dan
          sekaligus menghindari timeout request pada batch sangat besar.
        """
        if not self:
            raise UserError(_("Tidak ada baris untuk dicetak."))

        if len(self) <= CETAK_CHUNK_SIZE:
            return self.env.ref('ka_cetak.action_report_ka_cetak_ngp').report_action(self)

        job_id = self.env['ka.cetak.print.job'].create_job(self.ids, nama_file)
        return {
            'type': 'ir.actions.client',
            'tag': 'ka_cetak_print_progress',
            'target': 'new',
            'params': {
                'job_id': job_id,
                'total': len(self),
            },
        }

    def action_cetak_terpilih(self):
        """Dipakai dari tombol/aksi di list view untuk mencetak baris terpilih."""
        lines = self.filtered('is_valid')
        if not lines:
            raise UserError(_(
                "Tidak ada baris valid pada pilihan. Pastikan register sudah "
                "divalidasi dan ditemukan di master."))
        return lines._cetak_ngp_pdf(nama_file='NGP_terpilih')

