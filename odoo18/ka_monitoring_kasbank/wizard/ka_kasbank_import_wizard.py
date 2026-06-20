# -*- coding: utf-8 -*-
import base64
import io
import re
import datetime
import logging

from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

DEFAULT_SHEET = {
    'balance': 'KASBANK',
    'inventory': 'Gula Eks 2026',
    'sales_bulk': 'Penjualan Bulk',
    'sales_retail': 'Penjualan Ritel',
}


class KaKasbankImportWizard(models.TransientModel):
    _name = 'ka.kasbank.import.wizard'
    _description = 'Wizard Import Data Kas/Bank & Persediaan'

    import_type = fields.Selection(
        [('balance', 'Saldo Kas/Bank (sheet KASBANK)'),
         ('inventory', 'Persediaan (sheet Gula Eks ...)'),
         ('sales_bulk', 'Penjualan Bulk (sheet Penjualan Bulk)'),
         ('sales_retail', 'Penjualan Ritel (sheet Penjualan Ritel)')],
        string='Jenis Data', required=True, default='balance')
    file = fields.Binary(string='File Excel (.xlsx)', required=True)
    filename = fields.Char(string='Nama File')
    sheet_name = fields.Char(
        string='Nama Sheet',
        help='Kosongkan untuk default per jenis. Untuk Persediaan, ganti ke '
             '"Gula Eks 2025" / "Gula Eks 2026" sesuai tahun.')
    company_id = fields.Many2one(
        'res.company', string='Unit/Company', required=True,
        default=lambda self: self.env.company,
        help='Untuk Penjualan Bulk diabaikan — unit ditentukan otomatis per blok (KBA/Trangkil).')
    update_existing = fields.Boolean(
        string='Perbarui yang sudah ada', default=False,
        help='Jika data dengan kunci sama sudah ada: dicentang → diperbarui; '
             'tidak → dilewati.')
    create_missing = fields.Boolean(
        string='Buat master/partner yang belum ada', default=True)

    state = fields.Selection([('draft', 'Draft'), ('done', 'Selesai')],
                             default='draft')
    result_summary = fields.Text(string='Ringkasan', readonly=True)

    # ───────────────────────── helper sel ─────────────────────────
    @staticmethod
    def _cs(v):
        return str(v).strip() if v not in (None, '') else ''

    @staticmethod
    def _cf(v):
        return float(v) if isinstance(v, (int, float)) else 0.0

    @staticmethod
    def _cd(v):
        if isinstance(v, datetime.datetime):
            return v.date()
        if isinstance(v, datetime.date):
            return v
        return None

    def _get_company(self, name):
        if name:
            comp = self.env['res.company'].search([('name', '=ilike', name)], limit=1)
            if comp:
                return comp
        return self.company_id

    def _get_partner(self, name):
        if not name:
            return False
        p = self.env['res.partner'].search([('name', '=ilike', name)], limit=1)
        if not p and self.create_missing:
            p = self.env['res.partner'].create({'name': name})
        return p.id if p else False

    def _get_account(self, name):
        if not name:
            return False
        acc = self.env['ka.kasbank.account'].search([('name', '=ilike', name)], limit=1)
        if not acc and self.create_missing:
            acc = self.env['ka.kasbank.account'].create({
                'name': name,
                'account_type': 'cash' if 'kas' in name.lower() else 'bank',
                'company_id': self.company_id.id,
            })
        return acc

    def _get_product(self, name):
        if not name:
            return False
        pr = self.env['ka.kasbank.product'].search([('name', '=ilike', name)], limit=1)
        if not pr and self.create_missing:
            n = name.lower()
            cat = 'gkp'
            if 'retail' in n or 'ritel' in n:
                cat = 'ritel'
            elif 'setengah' in n:
                cat = 'setengah_jadi'
            elif 'sip' in n:
                cat = 'sip'
            pr = self.env['ka.kasbank.product'].create({
                'name': name, 'category': cat,
                'satuan': 'kg' if cat == 'ritel' else 'ton',
            })
        return pr

    # ───────────────────────── entry point ─────────────────────────
    def action_import(self):
        self.ensure_one()
        try:
            import openpyxl  # noqa
        except ImportError:
            raise UserError(_(
                "Library 'openpyxl' belum terpasang di server Odoo.\n"
                "Pasang dengan: pip install openpyxl"))

        if not self.file:
            raise UserError(_("Unggah file Excel terlebih dahulu."))

        # nonaktifkan tracking saat import massal (hindari ribuan pesan chatter)
        self = self.with_context(
            tracking_disable=True, mail_create_nolog=True, mail_notrack=True)

        try:
            data = base64.b64decode(self.file)
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        except Exception as e:
            raise UserError(_("Gagal membuka file Excel: %s") % e)

        sheet = self.sheet_name or DEFAULT_SHEET.get(self.import_type)
        if sheet not in wb.sheetnames:
            raise UserError(_(
                "Sheet '%(s)s' tidak ditemukan.\nSheet tersedia: %(list)s",
                s=sheet, list=", ".join(wb.sheetnames)))
        ws = wb[sheet]

        if self.import_type == 'balance':
            c, u, s = self._import_balance(ws)
            label = _("Saldo Kas/Bank")
        elif self.import_type == 'inventory':
            m = re.search(r'(20\d{2})', sheet)
            year = int(m.group(1)) if m else fields.Date.context_today(self).year
            c, u, s = self._import_inventory(ws, year)
            label = _("Persediaan (Eks %s)") % year
        elif self.import_type == 'sales_bulk':
            c, u, s = self._import_bulk(ws)
            label = _("Penjualan Bulk")
        else:
            c, u, s = self._import_retail(ws)
            label = _("Penjualan Ritel")

        summary = _(
            "Import %(label)s dari sheet '%(sheet)s':\n"
            "  • Dibuat   : %(c)s\n"
            "  • Diperbarui: %(u)s\n"
            "  • Dilewati  : %(s)s",
            label=label, sheet=sheet, c=c, u=u, s=s)
        self.write({'state': 'done', 'result_summary': summary})
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
            'context': self.env.context,
        }

    # ───────────────────────── parsers ─────────────────────────
    def _import_balance(self, ws):
        Balance = self.env['ka.kasbank.balance']
        Line = self.env['ka.kasbank.balance.line']
        accounts = {}
        for c in range(2, 7):
            h = self._cs(ws.cell(3, c).value)
            if h and h.lower() != 'total':
                accounts[c] = h
        created = updated = skipped = 0
        for r in range(4, ws.max_row + 1):
            d = self._cd(ws.cell(r, 1).value)
            if not d:
                continue
            pairs = [(accounts[c], self._cf(ws.cell(r, c).value))
                     for c in accounts if ws.cell(r, c).value not in (None, '')]
            if not pairs:
                continue
            existing = Balance.search(
                [('date', '=', d), ('company_id', '=', self.company_id.id)], limit=1)
            if existing and not self.update_existing:
                skipped += 1
                continue
            if existing:
                snap = existing
                updated += 1
            else:
                snap = Balance.create({'date': d, 'company_id': self.company_id.id})
                snap.line_ids.unlink()  # buang baris nol auto-populate
                created += 1
            for name, bal in pairs:
                acc = self._get_account(name)
                if not acc:
                    continue
                line = snap.line_ids.filtered(lambda l: l.account_id.id == acc.id)
                if line:
                    line.balance = bal
                else:
                    Line.create({'balance_id': snap.id, 'account_id': acc.id, 'balance': bal})
        return created, updated, skipped

    def _import_inventory(self, ws, year):
        Inv = self.env['ka.kasbank.inventory']
        created = updated = skipped = 0
        seen = set()  # blok pertama per produk yang menang (sheet kadang punya blok kembar kosong)
        c = 1
        while c <= ws.max_column:
            product_name = self._cs(ws.cell(4, c).value)
            key_name = product_name.lower()
            if product_name and key_name not in seen:
                seen.add(key_name)
                product = self._get_product(product_name)
                if product:
                    for r in range(7, ws.max_row + 1):
                        d = self._cd(ws.cell(r, c).value)
                        if not d:
                            continue
                        vals = {
                            'saldo_awal': self._cf(ws.cell(r, c + 1).value),
                            'produksi': self._cf(ws.cell(r, c + 2).value),
                            'penjualan': self._cf(ws.cell(r, c + 3).value),
                        }
                        existing = Inv.search([
                            ('date', '=', d), ('company_id', '=', self.company_id.id),
                            ('product_id', '=', product.id), ('production_year', '=', year),
                        ], limit=1)
                        if existing:
                            if self.update_existing:
                                existing.write(vals)
                                updated += 1
                            else:
                                skipped += 1
                        else:
                            Inv.create(dict(vals, date=d, company_id=self.company_id.id,
                                            product_id=product.id, production_year=year))
                            created += 1
            c += 6
        return created, updated, skipped

    def _import_bulk(self, ws):
        Bulk = self.env['ka.kasbank.sales.bulk']
        created = updated = skipped = 0
        for start, label in [(1, 1), (13, 13)]:
            company = self._get_company(self._cs(ws.cell(3, label).value))
            for r in range(6, ws.max_row + 1):
                sp = self._cs(ws.cell(r, start).value)
                if not sp:
                    continue
                pay_raw = ws.cell(r, start + 7).value
                pay = (self._cd(pay_raw).strftime('%d/%m/%Y')
                       if self._cd(pay_raw) else self._cs(pay_raw))
                status = ('lunas' if self._cs(ws.cell(r, start + 8).value).lower() == 'lunas'
                          else 'open')
                pname = self._cs(ws.cell(r, start + 9).value)
                product = self._get_product(pname) if pname else False
                vals = {
                    'date': self._cd(ws.cell(r, start + 1).value),
                    'partner_id': self._get_partner(self._cs(ws.cell(r, start + 2).value)),
                    'production_year': int(self._cf(ws.cell(r, start + 3).value)) or False,
                    'quantity': self._cf(ws.cell(r, start + 4).value),
                    'price_unit': self._cf(ws.cell(r, start + 5).value),
                    'payment_date': pay,
                    'state': status,
                    'product_id': product.id if product else False,
                }
                existing = Bulk.search(
                    [('sp_number', '=', sp), ('company_id', '=', company.id)], limit=1)
                if existing:
                    if self.update_existing:
                        existing.write(vals)
                        updated += 1
                    else:
                        skipped += 1
                else:
                    Bulk.create(dict(vals, sp_number=sp, company_id=company.id))
                    created += 1
        return created, updated, skipped

    def _import_retail(self, ws):
        Ret = self.env['ka.kasbank.sales.retail']
        seg_keys = dict(Ret._fields['segment'].selection)
        created = updated = skipped = 0
        for r in range(6, ws.max_row + 1):
            code = self._cs(ws.cell(r, 2).value)
            inv = self._cs(ws.cell(r, 8).value)
            if not (code or inv):
                continue
            seg = self._cs(ws.cell(r, 4).value).upper()
            seg = seg if seg in seg_keys else False
            status_raw = self._cs(ws.cell(r, 14).value).lower()
            status = {'paid': 'paid', 'overdue': 'overdue'}.get(status_raw, 'open')
            vals = {
                'customer_code': code,
                'partner_id': self._get_partner(self._cs(ws.cell(r, 3).value)),
                'segment': seg,
                'so_number': self._cs(ws.cell(r, 5).value),
                'delivery_order': self._cs(ws.cell(r, 6).value),
                'driver': self._cs(ws.cell(r, 7).value),
                'invoice_date': self._cd(ws.cell(r, 9).value),
                'due_date': self._cd(ws.cell(r, 10).value),
                'qty_kg': self._cf(ws.cell(r, 11).value),
                'price_unit': self._cf(ws.cell(r, 12).value),
                'payment_status': status,
                'amount_paid': self._cf(ws.cell(r, 15).value),
                'paid_month': self._cs(ws.cell(r, 16).value),
                'company_id': self.company_id.id,
            }
            existing = (Ret.search([('invoice_number', '=', inv),
                                    ('company_id', '=', self.company_id.id)], limit=1)
                        if inv else False)
            if existing:
                if self.update_existing:
                    existing.write(vals)
                    updated += 1
                else:
                    skipped += 1
            else:
                Ret.create(dict(vals, invoice_number=inv))
                created += 1
        return created, updated, skipped
