# -*- coding: utf-8 -*-
import base64
import io
import math
import logging
from odoo import models, fields, api, _
from odoo.exceptions import UserError


def _round0(value):
    """Pembulatan ke integer (ROUND_HALF_UP)."""
    from decimal import Decimal, ROUND_HALF_UP
    if not value:
        return 0
    return int(Decimal(str(value)).quantize(Decimal('1'), rounding=ROUND_HALF_UP))

_logger = logging.getLogger(__name__)


class KaKetentuan(models.Model):
    """
    Master Ketentuan — harga, faktor rendemen, bagi hasil.
    Per-periode. Yang terakhir dibuat menang saat overlap.
    """
    _name = 'ka.ketentuan'
    _description = 'Ketentuan Bagi Hasil'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _rec_name = 'name'
    _order = 'create_date DESC'

    name = fields.Char(
        string='No. Ketentuan', readonly=True, copy=False,
        default='/', index=True
    )
    company_id = fields.Many2one(
        'res.company', string='Unit/Company', required=True,
        default=lambda self: self.env.company, index=True
    )
    tanggal = fields.Date(
        string='Tanggal Dibuat', required=True,
        default=fields.Date.today, tracking=True
    )
    tgl_berlaku = fields.Datetime(
        string='Berlaku Hingga', required=True, tracking=True,
        help='Ketentuan berlaku untuk timbang dengan waktu <= waktu ini (default jam 06:00:00)'
    )

    # ── Harga ──────────────────────────────────────────────────
    harga_gula  = fields.Float(string='Harga Gula', digits=(16, 2), tracking=True)
    harga_tetes = fields.Float(string='Harga Tetes', digits=(16, 2), tracking=True)

    # ── Faktor Rendemen ────────────────────────────────────────
    faktor_gula     = fields.Float(string='Faktor Rendemen Gula', digits=(10, 4), tracking=True)
    faktor_tetes    = fields.Float(string='Faktor Rendemen Tetes', digits=(10, 4), tracking=True)
    faktor_bh_tetes = fields.Float(string='Faktor B.H Tetes', digits=(10, 4), tracking=True)
    faktor_nira     = fields.Float(string='Faktor Rendemen Nira', digits=(10, 4), tracking=True)

    # ── Pembagian Gula BH (jual : natura) ──────────────────────
    persen_lelang = fields.Float(
        string='% Gula Jual (Lelang)', digits=(5, 2), default=80.0, tracking=True,
        help='Persentase bagian gula jual/lelang dari Gula BH.'
    )
    persen_natura = fields.Float(
        string='% Natura (Gula Kecil)', digits=(5, 2), default=20.0, tracking=True,
        help='Persentase bagian natura/gula kecil dari Gula BH. '
             'natura_kecil = floor0(gula_bh × %natura), gula_jual = gula_bh − natura_kecil.'
    )

    # ── Titipan Tetes (Natura3 / sharing tetes) ────────────────
    titipan_tetes = fields.Float(
        string='Titipan Tetes', digits=(10, 4), tracking=True,
        help='Nilai titipan tetes per kuintal untuk Natura3. '
             'natura3 = floor0(titipan_tetes × netto_relaksasi).'
    )

    # ── Bagi Hasil ─────────────────────────────────────────────
    bagi_hasil_default = fields.Float(
        string='Bagi Hasil Default', digits=(10, 4), tracking=True,
        help='Dipakai untuk register yang tidak ada di list bagi hasil.'
    )

    # ── Gula Kawalan ───────────────────────────────────────────
    gula_kawalan_default = fields.Float(
        string='Gula Kawalan Default', digits=(10, 4), tracking=True,
        help='Nilai kawalan yang diberikan ke register yang ada di daftar '
             'kawalan. Register di luar daftar tidak mendapatkan kawalan.'
    )

    keterangan = fields.Text(string='Keterangan')

    state = fields.Selection([
        ('active', 'Aktif'),
        ('cancel', 'Dibatalkan'),
    ], string='Status', default='active', tracking=True)

    line_ids = fields.One2many(
        'ka.ketentuan.line', 'ketentuan_id', string='Bagi Hasil per Register'
    )
    line_count = fields.Integer(string='Jml Register', compute='_compute_line_count')

    kawalan_line_ids = fields.One2many(
        'ka.ketentuan.kawalan.line', 'ketentuan_id',
        string='Register Penerima Kawalan'
    )
    kawalan_count = fields.Integer(
        string='Jml Register Kawalan', compute='_compute_kawalan_count'
    )

    # ── Upload ─────────────────────────────────────────────────
    upload_file = fields.Binary(string='File Excel (.xlsx)')
    upload_filename = fields.Char(string='Nama File')

    upload_kawalan_file = fields.Binary(string='File Excel Kawalan (.xlsx)')
    upload_kawalan_filename = fields.Char(string='Nama File Kawalan')

    @api.depends('line_ids')
    def _compute_line_count(self):
        for rec in self:
            rec.line_count = len(rec.line_ids)

    @api.depends('kawalan_line_ids')
    def _compute_kawalan_count(self):
        for rec in self:
            rec.kawalan_count = len(rec.kawalan_line_ids)

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if vals.get('name', '/') == '/':
                vals['name'] = self.env['ir.sequence'].next_by_code('ka.ketentuan') or '/'
        return super().create(vals_list)

    def action_cancel(self):
        self.ensure_one()
        if not (
            self.env.user.has_group('ka_user_management.group_ka_closing') or
            self.env.user.has_group('ka_user_management.group_ka_admin')
        ):
            raise UserError(_('Anda tidak memiliki akses untuk membatalkan ketentuan.'))
        self.write({'state': 'cancel'})

    def action_set_active(self):
        self.ensure_one()
        self.write({'state': 'active'})

    def unlink(self):
        raise UserError(_(
            'Record Ketentuan tidak dapat dihapus untuk menjaga history. '
            'Gunakan tombol Batalkan.'
        ))

    def action_import_bagihasil(self):
        """Import list bagi hasil per register dari Excel."""
        self.ensure_one()
        if not self.upload_file:
            raise UserError(_('Harap pilih file Excel terlebih dahulu.'))

        try:
            from openpyxl import load_workbook
        except ImportError:
            raise UserError(_('Library openpyxl belum terpasang di server. Hubungi Administrator.'))

        try:
            data = base64.b64decode(self.upload_file)
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            raise UserError(_('Gagal membaca file Excel: %s') % str(e))

        # Hapus line lama
        self.line_ids.unlink()

        Line = self.env['ka.ketentuan.line']
        vals_list = []
        skipped = 0
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or row[0] is None:
                continue
            register = str(row[0]).strip()
            if not register:
                continue
            try:
                bagi_hasil = float(row[1]) if row[1] is not None else 0.0
            except (ValueError, TypeError):
                skipped += 1
                continue
            vals_list.append({
                'ketentuan_id': self.id,
                'register': register,
                'bagi_hasil': bagi_hasil,
            })

        if vals_list:
            Line.create(vals_list)

        msg = f'{len(vals_list)} register berhasil diimpor.'
        if skipped:
            msg += f' {skipped} baris dilewati (format tidak valid).'

        self.write({'upload_file': False, 'upload_filename': False})

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Import Selesai',
                'message': msg,
                'type': 'success',
                'sticky': False,
            }
        }

    def action_import_kawalan(self):
        """Import daftar register penerima gula kawalan dari Excel (kolom: register)."""
        self.ensure_one()
        if not self.upload_kawalan_file:
            raise UserError(_('Harap pilih file Excel kawalan terlebih dahulu.'))

        try:
            from openpyxl import load_workbook
        except ImportError:
            raise UserError(_('Library openpyxl belum terpasang di server. Hubungi Administrator.'))

        try:
            data = base64.b64decode(self.upload_kawalan_file)
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            raise UserError(_('Gagal membaca file Excel: %s') % str(e))

        # Hapus daftar kawalan lama
        self.kawalan_line_ids.unlink()

        Line = self.env['ka.ketentuan.kawalan.line']
        vals_list = []
        seen = set()
        skipped = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            register = str(row[0]).strip()
            if not register:
                continue
            # Hindari duplikat register dalam satu file
            if register in seen:
                skipped += 1
                continue
            seen.add(register)
            vals_list.append({
                'ketentuan_id': self.id,
                'register': register,
            })

        if vals_list:
            Line.create(vals_list)

        msg = f'{len(vals_list)} register kawalan berhasil diimpor.'
        if skipped:
            msg += f' {skipped} baris dilewati (kosong/duplikat).'

        self.write({'upload_kawalan_file': False, 'upload_kawalan_filename': False})

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Import Kawalan Selesai',
                'message': msg,
                'type': 'success',
                'sticky': False,
            }
        }

    @api.model
    def get_ketentuan_aktif(self, waktu_timbang, company_id):
        """
        Ambil ketentuan yang berlaku untuk waktu timbang tertentu.
        Logika sama dengan relaksasi: filter tgl_berlaku, fallback terakhir dibuat.
        Return: record ka.ketentuan atau None.
        """
        domain_base = [
            ('company_id', '=', company_id),
            ('state', '=', 'active'),
        ]
        ket = self.sudo().search(
            domain_base + [('tgl_berlaku', '>=', waktu_timbang)],
            order='create_date DESC', limit=1
        )
        if not ket:
            ket = self.sudo().search(domain_base, order='create_date DESC', limit=1)
        return ket or None

    def get_bagi_hasil_for_register(self, register_code):
        """Ambil BH untuk register: dari list jika ada, else default."""
        self.ensure_one()
        line = self.line_ids.filtered(lambda l: l.register == register_code)
        if line:
            return line[0].bagi_hasil
        return self.bagi_hasil_default

    def get_kawalan_for_register(self, register_code):
        """
        Ambil nilai gula kawalan untuk register.
        Register HANYA dapat kawalan jika ada di daftar kawalan (hasil import).
        - Ada di daftar  → gula_kawalan_default
        - Tidak ada      → 0.0 (tidak dapat kawalan)
        """
        self.ensure_one()
        line = self.kawalan_line_ids.filtered(lambda l: l.register == register_code)
        if line:
            return self.gula_kawalan_default
        return 0.0

    @api.constrains('persen_lelang', 'persen_natura')
    def _check_persen_split(self):
        for rec in self:
            total = (rec.persen_lelang or 0.0) + (rec.persen_natura or 0.0)
            # Toleransi kecil untuk pembulatan desimal
            if abs(total - 100.0) > 0.01:
                raise UserError(_(
                    'Persentase Gula Jual (%.2f%%) + Natura (%.2f%%) harus = 100%%. '
                    'Saat ini: %.2f%%.'
                ) % (rec.persen_lelang or 0.0, rec.persen_natura or 0.0, total))

    # ─────────────────────────────────────────────────────────
    #  HELPER PERHITUNGAN (dipakai oleh NTP)
    #  Semua rumus terpusat di sini karena inputnya = field Ketentuan.
    # ─────────────────────────────────────────────────────────
    def compute_split_gula(self, gula_bh):
        """
        Bagi Gula BH (integer) menjadi (gula_jual, natura_kecil) sesuai FoxPro:
        natura_kecil = round(gula_bh × %natura)  [ROUND_HALF_UP], gula_jual = gula_bh − natura_kecil.
        """
        self.ensure_one()
        from decimal import Decimal, ROUND_HALF_UP
        gula_bh = int(gula_bh or 0)
        raw = gula_bh * (self.persen_natura or 0.0) / 100.0
        natura_kecil = int(Decimal(str(raw)).quantize(Decimal('1'), rounding=ROUND_HALF_UP)) if raw else 0
        gula_jual = gula_bh - natura_kecil
        return gula_jual, natura_kecil

    def compute_natura2_tetes(self, netto_relaksasi):
        """Natura2 (tetes berlaku) = round((harga_tetes ÷ harga_gula) × netto)."""
        self.ensure_one()
        if not self.harga_gula:
            return 0
        return _round0((self.harga_tetes / self.harga_gula) * (netto_relaksasi or 0.0))

    def compute_natura3_titipan(self, netto_relaksasi):
        """Natura3 (sharing tetes) = round(titipan_tetes × netto)."""
        self.ensure_one()
        return _round0((self.titipan_tetes or 0.0) * (netto_relaksasi or 0.0))

    def compute_tetes_tani(self, netto_relaksasi):
        """
        Tetes petani (teori) = floor2(netto × faktor_tetes). Disimpan, tidak ditampilkan.
        Return: (tetes_tani, nilai_tetes) di mana nilai_tetes = tetes_tani × harga_tetes.
        """
        self.ensure_one()
        raw = (self.faktor_tetes or 0.0) * (netto_relaksasi or 0.0)
        tetes_tani = math.floor(raw * 100) / 100.0
        nilai_tetes = tetes_tani * (self.harga_tetes or 0.0)
        return tetes_tani, nilai_tetes


class KaKetentuanLine(models.Model):
    """Bagi hasil per register (dari upload Excel)."""
    _name = 'ka.ketentuan.line'
    _description = 'Bagi Hasil per Register'
    _order = 'register'

    ketentuan_id = fields.Many2one(
        'ka.ketentuan', string='Ketentuan', required=True, ondelete='cascade'
    )
    company_id = fields.Many2one(
        'res.company', related='ketentuan_id.company_id',
        store=True, index=True
    )
    register = fields.Char(string='Register', required=True, index=True)
    bagi_hasil = fields.Float(string='Bagi Hasil', digits=(10, 4), required=True)


class KaKetentuanKawalanLine(models.Model):
    """Daftar register penerima gula kawalan (dari upload Excel, kolom: register)."""
    _name = 'ka.ketentuan.kawalan.line'
    _description = 'Register Penerima Gula Kawalan'
    _order = 'register'

    ketentuan_id = fields.Many2one(
        'ka.ketentuan', string='Ketentuan', required=True, ondelete='cascade'
    )
    company_id = fields.Many2one(
        'res.company', related='ketentuan_id.company_id',
        store=True, index=True
    )
    register = fields.Char(string='Register', required=True, index=True)
