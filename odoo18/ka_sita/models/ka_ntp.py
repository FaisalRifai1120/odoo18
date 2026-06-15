# -*- coding: utf-8 -*-
import base64
import io
import math
import logging
from datetime import datetime, time
from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


def round2(value):
    """Pembulatan 2 desimal (ROUND_HALF_UP). Contoh: 0.345 -> 0.35"""
    from decimal import Decimal, ROUND_HALF_UP
    if not value:
        return 0.0
    return float(Decimal(str(value)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))


def floor2(value):
    """Pembulatan ke bawah 2 desimal. Contoh: 0.058 -> 0.05"""
    if not value:
        return 0.0
    return math.floor(float(value) * 100) / 100.0


def floor0(value):
    """Pembulatan ke bawah ke integer. Contoh: 9950.7 -> 9950"""
    if not value:
        return 0
    return int(math.floor(float(value)))


def round0(value):
    """Pembulatan ke integer (ROUND_HALF_UP). Contoh: 89.5 -> 90, 89.4 -> 89"""
    from decimal import Decimal, ROUND_HALF_UP
    if not value:
        return 0
    return int(Decimal(str(value)).quantize(Decimal('1'), rounding=ROUND_HALF_UP))


class KaNtp(models.Model):
    """
    NTP — Nota Tebu Petani.
    Tahap 4a: rekap PER TRUK (belum agregat per register).
    """
    _name = 'ka.ntp'
    _description = 'Nota Tebu Petani'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _rec_name = 'name'
    _order = 'create_date DESC'

    name = fields.Char(
        string='No. NTP', readonly=True, copy=False,
        default='/', index=True
    )
    company_id = fields.Many2one(
        'res.company', string='Unit/Company', required=True,
        default=lambda self: self.env.company, index=True
    )
    tanggal_pembuatan = fields.Date(
        string='Tanggal Pembuatan', required=True,
        default=fields.Date.today, tracking=True
    )
    tgl_awal = fields.Datetime(
        string='Tanggal Awal', required=True, tracking=True,
        help='Default jam 06:00:00'
    )
    tgl_akhir = fields.Datetime(
        string='Tanggal Akhir', required=True, tracking=True,
        help='Default jam 05:59:59'
    )
    periode = fields.Char(string='Periode', tracking=True)

    ketentuan_id = fields.Many2one(
        'ka.ketentuan', string='Ketentuan',
        domain="[('state','=','active'),('company_id','=',company_id)]",
        tracking=True,
        help='Ketentuan yang dipakai untuk perhitungan. '
             'Kosongkan untuk auto-pilih ketentuan aktif per waktu timbang (estimasi). '
             'Pilih satu ketentuan untuk mengunci perhitungan saat closing.'
    )

    # ── Header Kwitansi ────────────────────────────────────────
    kode_periode = fields.Char(
        string='Kode Periode', size=3,
        help='Akhiran NOMOR kwitansi, mis. "A" pada "00003 / A".'
    )
    kwitansi_kota = fields.Char(string='Kota Kwitansi', default='Malang')
    pejabat_jabatan = fields.Char(string='Jabatan Penandatangan', default='Kepala Bag. TUK')
    pejabat_nama = fields.Char(string='Nama Penandatangan')

    jenis_kelompok = fields.Selection([
        ('jenis_register', 'Jenis Register (TR/TS)'),
        ('metode',         'Metode (SBH/SPT)'),
        ('digit',          'Per Digit Register'),
    ], string='Jenis Pengelompokan', default='jenis_register', tracking=True)

    # Filter opsional berdasarkan jenis
    filter_jenis_register = fields.Selection([
        ('TR', 'TR'),
        ('TS', 'TS'),
    ], string='Filter Jenis Register')
    filter_metode = fields.Selection([
        ('SBH', 'SBH'),
        ('SPT', 'SPT'),
    ], string='Filter Metode')
    filter_digit_posisi = fields.Integer(string='Posisi Digit Filter')
    filter_digit_nilai = fields.Char(string='Nilai Digit Filter', size=2)

    state = fields.Selection([
        ('draft',  'Draft'),
        ('proses', 'Proses'),
        ('done',   'Selesai'),
    ], string='Status', default='draft', tracking=True)

    line_ids = fields.One2many('ka.ntp.line', 'ntp_id', string='Detail per Truk')
    nota_gula_ids = fields.One2many('ka.ntp.nota.gula', 'ntp_id', string='Nota Gula')
    line_count = fields.Integer(string='Jml Truk', compute='_compute_totals')

    total_netto       = fields.Float(string='Total Netto', compute='_compute_totals', digits=(16, 2))
    total_netto_relaksasi = fields.Float(string='Total Tebu Final', compute='_compute_totals', digits=(16, 2))
    total_gula        = fields.Integer(string='Total Gula BH', compute='_compute_totals')
    total_rupiah      = fields.Float(string='Total Rupiah BH', compute='_compute_totals', digits=(16, 2))
    total_gula_kawalan = fields.Integer(string='Total Gula Kawalan', compute='_compute_totals')
    total_gula_total  = fields.Integer(string='Total Gula (Keseluruhan)', compute='_compute_totals')
    total_rupiah_total = fields.Float(string='Total Rupiah (Keseluruhan)', compute='_compute_totals', digits=(16, 2))

    # Status import (penanda override)
    has_import_relaksasi = fields.Boolean(string='Ada Import Relaksasi', default=False)
    has_import_bagihasil = fields.Boolean(string='Ada Import Bagi Hasil', default=False)
    has_import_kawalan = fields.Boolean(string='Ada Import Kawalan', default=False)

    # Upload fields
    upload_relaksasi_file = fields.Binary(string='File Relaksasi (.xlsx)')
    upload_relaksasi_filename = fields.Char(string='Nama File Relaksasi')
    upload_bagihasil_file = fields.Binary(string='File Bagi Hasil (.xlsx)')
    upload_bagihasil_filename = fields.Char(string='Nama File Bagi Hasil')
    upload_kawalan_file = fields.Binary(string='File Kawalan (.xlsx)')
    upload_kawalan_filename = fields.Char(string='Nama File Kawalan')

    # File import tersimpan (untuk download ulang)
    import_relaksasi_saved = fields.Binary(string='File Relaksasi Tersimpan', attachment=True)
    import_relaksasi_filename = fields.Char(string='Nama File Relaksasi Tersimpan')
    import_bagihasil_saved = fields.Binary(string='File Bagi Hasil Tersimpan', attachment=True)
    import_bagihasil_filename = fields.Char(string='Nama File Bagi Hasil Tersimpan')
    import_kawalan_saved = fields.Binary(string='File Kawalan Tersimpan', attachment=True)
    import_kawalan_filename = fields.Char(string='Nama File Kawalan Tersimpan')

    @api.depends('line_ids', 'line_ids.netto', 'line_ids.netto_relaksasi',
                 'line_ids.register', 'line_ids.bagi_hasil', 'line_ids.nilai_kawalan',
                 'line_ids.harga_gula', 'ketentuan_id')
    def _compute_totals(self):
        for rec in self:
            rec.line_count = len(rec.line_ids)
            rec.total_netto = sum(rec.line_ids.mapped('netto'))
            rec.total_netto_relaksasi = sum(rec.line_ids.mapped('netto_relaksasi'))
            # Gula & rupiah dihitung di level REGISTER (floor sekali) — match FoxPro
            rekap = rec._register_rekap()
            rec.total_gula = sum(d['gula'] for d in rekap)
            rec.total_rupiah = sum(d['rupiah'] for d in rekap)
            rec.total_gula_kawalan = sum(d['gula_kawalan'] for d in rekap)
            rec.total_gula_total = sum(d['gula_total'] for d in rekap)
            rec.total_rupiah_total = sum(d['rupiah_total'] for d in rekap)

    def _resolve_ketentuan(self):
        """Ketentuan yang dipakai: pilihan manual > auto-pilih aktif per tgl_akhir."""
        self.ensure_one()
        Ketentuan = self.env['ka.ketentuan']
        if self.ketentuan_id:
            return self.ketentuan_id
        if self.tgl_akhir and self.company_id:
            return Ketentuan.get_ketentuan_aktif(self.tgl_akhir, self.company_id.id)
        return Ketentuan.browse()

    def _register_rekap(self):
        """
        Hitung agregat PER REGISTER (level register, floor sekali) — sesuai FoxPro.
        Tebu dijumlah dulu, baru gula/natura dihitung sekali per register.
        Return: list of dict per register.
        """
        self.ensure_one()
        if not self.line_ids:
            return []
        ket = self._resolve_ketentuan()
        persen_natura = ket.persen_natura if ket else 0.0

        # Agregasi input per register
        agg = {}
        for line in self.line_ids:
            reg = line.register or '-'
            d = agg.get(reg)
            if not d:
                d = agg[reg] = {
                    'register': reg,
                    'nama_register': line.nama_register,
                    'petani_id': line.petani_id.id if line.petani_id else False,
                    'tebu_final': 0.0,
                    'bagi_hasil': line.bagi_hasil,       # sama dalam 1 register
                    'nilai_kawalan': line.nilai_kawalan, # sama dalam 1 register
                    'harga_gula': ket.harga_gula if ket else line.harga_gula,
                }
            d['tebu_final'] += line.netto_relaksasi

        result = []
        for reg, d in agg.items():
            tebu = d['tebu_final']
            harga_gula = d['harga_gula'] or 0.0
            # Gula BH (floor sekali, BH presisi 4 desimal)
            gula_bh = floor0(tebu * d['bagi_hasil'])
            # Pembagian jual : natura — natura = ROUND(gula_bh × %natura)
            natura_kecil = round0(gula_bh * persen_natura / 100.0)
            gula_jual = gula_bh - natura_kecil
            # Natura tambahan (semua level register)
            gula_kawalan = round0(tebu * d['nilai_kawalan'])         # Natura1 = round(tebu × kg_ku)
            natura2 = ket.compute_natura2_tetes(tebu) if ket else 0
            natura3 = ket.compute_natura3_titipan(tebu) if ket else 0
            gula_total = gula_jual + natura_kecil + gula_kawalan + natura2 + natura3
            d.update({
                'gula': gula_bh,
                'gula_jual': gula_jual,
                'natura_kecil': natura_kecil,
                'gula_kawalan': gula_kawalan,
                'natura2': natura2,
                'natura3': natura3,
                'gula_total': gula_total,
                'rupiah': floor2(gula_bh * harga_gula),
                'rupiah_total': floor2(gula_total * harga_gula),
            })
            result.append(d)
        return result

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if vals.get('name', '/') == '/':
                vals['name'] = self.env['ir.sequence'].next_by_code('ka.ntp') or '/'
            # Default jam awal 06:00:00 dan akhir 05:59:59 jika hanya tanggal
            if vals.get('tgl_awal') and isinstance(vals['tgl_awal'], str) and len(vals['tgl_awal']) == 10:
                vals['tgl_awal'] = vals['tgl_awal'] + ' 06:00:00'
        return super().create(vals_list)

    # ─────────────────────────────────────────────────────────
    #  TOMBOL: REPROSES
    # ─────────────────────────────────────────────────────────
    def action_reproses(self):
        """Tombol Reproses: validasi + panggil _reproses_lines + notif."""
        self.ensure_one()
        if self.state == 'done':
            raise UserError(_('NTP yang sudah Selesai tidak bisa di-reproses.'))
        if not self.tgl_awal or not self.tgl_akhir:
            raise UserError(_('Harap isi Tanggal Awal dan Tanggal Akhir.'))
        count = self._reproses_lines()
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Reproses Selesai',
                'message': f'{count} data truk berhasil diproses.',
                'type': 'success',
                'sticky': False,
            }
        }

    def _reproses_lines(self):
        """
        Tarik ulang data dari ka.timbang.tebu (rentang tgl), hitung per truk:
        rafaksi_relaksasi, netto_relaksasi, gula, rupiah.
        Sumber aturan: import (jika ada), else dari menu Relaksasi & Ketentuan.
        Return: jumlah baris diproses.
        """
        self.ensure_one()
        if not self.tgl_awal or not self.tgl_akhir:
            return 0
        company_id = self.company_id.id
        ctx = dict(tracking_disable=True, mail_notrack=True, mail_create_nolog=True)

        # 1. Ambil data timbang dalam rentang
        Timbang = self.env['ka.timbang.tebu'].sudo()
        domain = [
            ('company_id', '=', company_id),
            ('date_out', '>=', self.tgl_awal),
            ('date_out', '<=', self.tgl_akhir),
        ]
        # Filter opsional berdasarkan jenis
        if self.jenis_kelompok == 'jenis_register' and self.filter_jenis_register:
            domain.append(('register_id.jenis_register', '=', self.filter_jenis_register))
        elif self.jenis_kelompok == 'metode' and self.filter_metode:
            domain.append(('register_id.metode', '=', self.filter_metode))

        timbang_recs = Timbang.search(domain)

        # Filter per-digit (dilakukan di Python karena karakter ke-N)
        if self.jenis_kelompok == 'digit' and self.filter_digit_posisi and self.filter_digit_nilai:
            pos = self.filter_digit_posisi
            nilai = str(self.filter_digit_nilai)
            timbang_recs = timbang_recs.filtered(
                lambda t: t.register and len(t.register) >= pos
                and str(t.register[pos - 1]) == nilai
            )

        if not timbang_recs:
            self.line_ids.unlink()
            return 0

        # 2. Ambil ketentuan & relaksasi aktif (untuk fallback non-import)
        Ketentuan = self.env['ka.ketentuan']
        Relaksasi = self.env['ka.relaksasi']

        # Hapus lines lama
        self.line_ids.unlink()

        # 3. Cache import override (kalau ada)
        import_relaksasi = {}  # register -> persentase
        import_bagihasil = {}  # register -> bagi_hasil
        import_kawalan = {}    # register -> nilai_kawalan
        if self.has_import_relaksasi:
            for l in self.env['ka.ntp.import.relaksasi'].search([('ntp_id', '=', self.id)]):
                import_relaksasi[l.register] = l.persentase
        if self.has_import_bagihasil:
            for l in self.env['ka.ntp.import.bagihasil'].search([('ntp_id', '=', self.id)]):
                import_bagihasil[l.register] = l.bagi_hasil
        if self.has_import_kawalan:
            for l in self.env['ka.ntp.import.kawalan'].search([('ntp_id', '=', self.id)]):
                import_kawalan[l.register] = l.kawalan

        Line = self.env['ka.ntp.line'].with_context(**ctx)
        vals_list = []

        for t in timbang_recs:
            register_code = t.register or ''
            waktu = t.date_out or self.tgl_akhir
            bruto = t.weight_kw or 0.0          # weight_kw = bruto (Netto Kw sebelum rafaksi)
            bobot_tebu = t.bobot_tebu or 0.0    # bobot_tebu = bruto - rafaksi asli
            rafaksi_asli = t.rafaksi or 0.0

            # ── Relaksasi ──
            persen_relaksasi = None
            if register_code in import_relaksasi:
                persen_relaksasi = import_relaksasi[register_code]
            else:
                persen_relaksasi = Relaksasi.get_relaksasi_for_register(
                    register_code, waktu, company_id
                )

            if rafaksi_asli > 0 and persen_relaksasi is not None:
                # rafaksi_relaksasi = round(rafaksi_asli x persentase, 2)
                rafaksi_relaksasi = round2(rafaksi_asli * (persen_relaksasi / 100.0))
                # tebu_final = bobot_tebu + rafaksi_relaksasi
                netto_relaksasi = round2(bobot_tebu + rafaksi_relaksasi)
            else:
                # tidak ada relaksasi → tebu final = bobot_tebu (bruto - rafaksi asli)
                rafaksi_relaksasi = 0.0
                netto_relaksasi = bobot_tebu if rafaksi_asli > 0 else bruto

            # ── Resolusi Ketentuan ──
            # Jika ketentuan_id dipilih → pakai itu untuk semua truk (closing).
            # Jika kosong → auto-pilih ketentuan aktif per waktu timbang (estimasi).
            ket = self.ketentuan_id or Ketentuan.get_ketentuan_aktif(waktu, company_id)

            # ── Bagi Hasil ──
            if register_code in import_bagihasil:
                bagi_hasil = import_bagihasil[register_code]
            else:
                bagi_hasil = ket.get_bagi_hasil_for_register(register_code) if ket else 0.0

            # Harga gula dari ketentuan terpilih/aktif
            harga_gula = ket.harga_gula if ket else 0.0

            # ── Nilai Kawalan (override import > ketentuan > 0) ──
            if register_code in import_kawalan:
                nilai_kawalan = import_kawalan[register_code]
            else:
                nilai_kawalan = ket.get_kawalan_for_register(register_code) if ket else 0.0

            # ── Gula BH & pembagian jual : natura ──
            gula = floor0(netto_relaksasi * bagi_hasil)       # Gula BH (integer)
            if ket:
                gula_jual, natura_kecil = ket.compute_split_gula(gula)
                natura2 = ket.compute_natura2_tetes(netto_relaksasi)
                natura3 = ket.compute_natura3_titipan(netto_relaksasi)
                tetes_tani, nilai_tetes = ket.compute_tetes_tani(netto_relaksasi)
            else:
                gula_jual, natura_kecil = gula, 0
                natura2 = 0
                natura3 = 0
                tetes_tani, nilai_tetes = 0.0, 0.0

            # ── Natura1 (gula kawalan) ──
            gula_kawalan = round0(netto_relaksasi * nilai_kawalan)

            # ── Total ──
            gula_total = gula_jual + natura_kecil + gula_kawalan + natura2 + natura3
            rupiah = floor2(gula * harga_gula)                # rupiah BH (kompatibilitas)
            rupiah_total = floor2(gula_total * harga_gula)    # rupiah keseluruhan

            vals_list.append({
                'ntp_id':            self.id,
                'timbang_id':        t.id,
                'no_timbang':        t.spta_id,
                'no_spta':           t.no_spta,
                'kd_antrian':        t.kd_antrian,
                'register':          register_code,
                'nama_register':     t.nama_register,
                'petani_id':         t.petani_id.id if t.petani_id else False,
                'date_out':          t.date_out,
                'netto':             bruto,
                'bobot_tebu':        bobot_tebu,
                'rafaksi_asli':      rafaksi_asli,
                'persen_relaksasi':  persen_relaksasi or 0.0,
                'rafaksi_relaksasi': rafaksi_relaksasi,
                'netto_relaksasi':   netto_relaksasi,
                'bagi_hasil':        bagi_hasil,
                'harga_gula':        harga_gula,
                'gula':              gula,
                'rupiah':            rupiah,
                'gula_jual':         gula_jual,
                'natura_kecil':      natura_kecil,
                'nilai_kawalan':     nilai_kawalan,
                'gula_kawalan':      gula_kawalan,
                'natura2':           natura2,
                'natura3':           natura3,
                'tetes_tani':        tetes_tani,
                'nilai_tetes':       nilai_tetes,
                'gula_total':        gula_total,
                'rupiah_total':      rupiah_total,
            })

        Line.create(vals_list)
        _logger.info('[KA-NTP] Reproses %s: %d truk diproses.', self.name, len(vals_list))
        return len(vals_list)

    # ─────────────────────────────────────────────────────────
    #  TOMBOL: IMPORT RELAKSASI
    # ─────────────────────────────────────────────────────────
    def action_import_relaksasi(self):
        self.ensure_one()
        if not self.upload_relaksasi_file:
            raise UserError(_('Harap pilih file Excel relaksasi terlebih dahulu.'))
        rows = self._read_xlsx(self.upload_relaksasi_file)

        self.env['ka.ntp.import.relaksasi'].search([('ntp_id', '=', self.id)]).unlink()
        vals_list = []
        for register, nilai in rows:
            vals_list.append({
                'ntp_id': self.id,
                'register': register,
                'persentase': nilai,
            })
        if vals_list:
            self.env['ka.ntp.import.relaksasi'].create(vals_list)

        self.write({
            'has_import_relaksasi': True,
            'import_relaksasi_saved': self.upload_relaksasi_file,
            'import_relaksasi_filename': self.upload_relaksasi_filename or 'relaksasi.xlsx',
            'upload_relaksasi_file': False,
            'upload_relaksasi_filename': False,
            'state': 'proses',
        })
        return self._notif(f'{len(vals_list)} relaksasi diimpor. Klik Reproses untuk menerapkan.')

    # ─────────────────────────────────────────────────────────
    #  TOMBOL: IMPORT BAGI HASIL
    # ─────────────────────────────────────────────────────────
    def action_import_bagihasil(self):
        self.ensure_one()
        if not self.upload_bagihasil_file:
            raise UserError(_('Harap pilih file Excel bagi hasil terlebih dahulu.'))
        rows = self._read_xlsx(self.upload_bagihasil_file)

        self.env['ka.ntp.import.bagihasil'].search([('ntp_id', '=', self.id)]).unlink()
        vals_list = []
        for register, nilai in rows:
            vals_list.append({
                'ntp_id': self.id,
                'register': register,
                'bagi_hasil': nilai,
            })
        if vals_list:
            self.env['ka.ntp.import.bagihasil'].create(vals_list)

        self.write({
            'has_import_bagihasil': True,
            'import_bagihasil_saved': self.upload_bagihasil_file,
            'import_bagihasil_filename': self.upload_bagihasil_filename or 'bagihasil.xlsx',
            'upload_bagihasil_file': False,
            'upload_bagihasil_filename': False,
            'state': 'proses',
        })
        return self._notif(f'{len(vals_list)} bagi hasil diimpor. Klik Reproses untuk menerapkan.')

    # ─────────────────────────────────────────────────────────
    #  TOMBOL: IMPORT KAWALAN
    # ─────────────────────────────────────────────────────────
    def action_import_kawalan(self):
        self.ensure_one()
        if not self.upload_kawalan_file:
            raise UserError(_('Harap pilih file Excel kawalan terlebih dahulu.'))
        rows = self._read_xlsx(self.upload_kawalan_file)

        self.env['ka.ntp.import.kawalan'].search([('ntp_id', '=', self.id)]).unlink()
        vals_list = []
        for register, nilai in rows:
            vals_list.append({
                'ntp_id': self.id,
                'register': register,
                'kawalan': nilai,
            })
        if vals_list:
            self.env['ka.ntp.import.kawalan'].create(vals_list)

        self.write({
            'has_import_kawalan': True,
            'import_kawalan_saved': self.upload_kawalan_file,
            'import_kawalan_filename': self.upload_kawalan_filename or 'kawalan.xlsx',
            'upload_kawalan_file': False,
            'upload_kawalan_filename': False,
            'state': 'proses',
        })
        return self._notif(f'{len(vals_list)} kawalan diimpor. Klik Reproses untuk menerapkan.')

    def _read_xlsx(self, file_b64):
        """Baca xlsx → list of (register, nilai_float). Skip header baris 1."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise UserError(_('Library openpyxl belum terpasang di server.'))
        try:
            data = base64.b64decode(file_b64)
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            raise UserError(_('Gagal membaca file Excel: %s') % str(e))

        result = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            register = str(row[0]).strip()
            if not register:
                continue
            try:
                nilai = float(row[1]) if row[1] is not None else 0.0
            except (ValueError, TypeError):
                continue
            result.append((register, nilai))
        return result

    def _notif(self, msg, kind='success'):
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {'title': 'Info', 'message': msg, 'type': kind, 'sticky': False},
        }

    @api.model
    def cron_auto_reproses(self):
        """
        Auto-reproses NTP draft yang periodenya masih aktif.
        Dipanggil sequential setelah cron sync timbang + QC.
        Hanya state draft + tgl_akhir >= sekarang.
        """
        now = fields.Datetime.now()
        ntps = self.search([
            ('state', '=', 'draft'),
            ('tgl_akhir', '>=', now),
        ])
        if not ntps:
            return
        _logger.info('[KA-NTP] Auto-reproses %d NTP draft aktif...', len(ntps))
        for ntp in ntps:
            try:
                with self.env.cr.savepoint():
                    ntp._reproses_lines()
                _logger.info('[KA-NTP] ✓ Auto-reproses %s: %d truk', ntp.name, len(ntp.line_ids))
            except Exception as e:
                _logger.error('[KA-NTP] ✗ Auto-reproses %s gagal: %s', ntp.name, str(e))

    def action_download_import_relaksasi(self):
        """Download file import relaksasi yang tersimpan."""
        self.ensure_one()
        if not self.import_relaksasi_saved:
            raise UserError(_('Tidak ada file import relaksasi tersimpan.'))
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/ka.ntp/{self.id}/import_relaksasi_saved/{self.import_relaksasi_filename or "relaksasi.xlsx"}?download=true',
            'target': 'self',
        }

    def action_download_import_bagihasil(self):
        """Download file import bagi hasil yang tersimpan."""
        self.ensure_one()
        if not self.import_bagihasil_saved:
            raise UserError(_('Tidak ada file import bagi hasil tersimpan.'))
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/ka.ntp/{self.id}/import_bagihasil_saved/{self.import_bagihasil_filename or "bagihasil.xlsx"}?download=true',
            'target': 'self',
        }

    def action_download_import_kawalan(self):
        """Download file import kawalan yang tersimpan."""
        self.ensure_one()
        if not self.import_kawalan_saved:
            raise UserError(_('Tidak ada file import kawalan tersimpan.'))
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/ka.ntp/{self.id}/import_kawalan_saved/{self.import_kawalan_filename or "kawalan.xlsx"}?download=true',
            'target': 'self',
        }

    def action_set_done(self):
        self.ensure_one()
        if not self.env.user.has_group('ka_user_management.group_ka_admin'):
            raise UserError(_('Hanya Administrator yang dapat menyelesaikan NTP.'))
        self.write({'state': 'done'})
        self._generate_nota_gula()

    def action_set_draft(self):
        self.ensure_one()
        # Hapus nota gula saat kembali ke draft
        self.nota_gula_ids.unlink()
        self.write({'state': 'draft'})

    def _generate_nota_gula(self):
        """Nota Gula per register — dihitung di level register (match FoxPro)."""
        self.ensure_one()
        self.nota_gula_ids.unlink()

        rekap = self._register_rekap()
        NotaGula = self.env['ka.ntp.nota.gula']
        vals_list = []
        for d in rekap:
            vals_list.append({
                'ntp_id': self.id,
                'register': d['register'],
                'nama_register': d['nama_register'],
                'petani_id': d['petani_id'],
                'tebu_final': round2(d['tebu_final']),
                'bagi_hasil': d['bagi_hasil'],
                'gula': d['gula'],
                'rupiah': round2(d['rupiah']),
                'gula_jual': d['gula_jual'],
                'natura_kecil': d['natura_kecil'],
                'gula_kawalan': d['gula_kawalan'],
                'natura2': d['natura2'],
                'natura3': d['natura3'],
                'gula_total': d['gula_total'],
                'rupiah_total': round2(d['rupiah_total']),
            })
        if vals_list:
            NotaGula.create(vals_list)

    def action_view_nota_gula(self):
        """Tampilkan Nota Gula (agregat per register). Hanya saat Done."""
        self.ensure_one()
        if self.state != 'done':
            raise UserError(_('Nota Gula hanya tersedia saat NTP berstatus Selesai.'))
        # Generate ulang agregat (pastikan fresh)
        self._generate_nota_gula()
        return {
            'type': 'ir.actions.act_window',
            'name': f'Nota Gula — {self.name}',
            'res_model': 'ka.ntp.nota.gula',
            'view_mode': 'list',
            'domain': [('ntp_id', '=', self.id)],
            'context': {'search_default_ntp': self.id},
            'target': 'current',
        }



class KaNtpLine(models.Model):
    """Detail NTP per truk."""
    _name = 'ka.ntp.line'
    _description = 'Detail NTP per Truk'
    _order = 'date_out, register'

    ntp_id = fields.Many2one('ka.ntp', string='NTP', required=True, ondelete='cascade')
    company_id = fields.Many2one('res.company', related='ntp_id.company_id', store=True, index=True)
    timbang_id = fields.Many2one('ka.timbang.tebu', string='Data Timbang', ondelete='set null')

    no_timbang = fields.Char(string='No. Timbang')
    no_spta    = fields.Char(string='No. SPTA')
    kd_antrian = fields.Char(string='No. Antrian')
    register   = fields.Char(string='Register', index=True)
    nama_register = fields.Char(string='Nama Register')
    petani_id  = fields.Many2one('ka.petani', string='Petani')
    date_out   = fields.Datetime(string='Tgl. Keluar')

    netto             = fields.Float(string='Bruto (Kw)', digits=(16, 2))
    bobot_tebu        = fields.Float(string='Bobot Tebu', digits=(16, 2))
    rafaksi_asli      = fields.Float(string='Rafaksi Asli', digits=(16, 2))
    persen_relaksasi  = fields.Float(string='Relaksasi (%)', digits=(5, 2))
    rafaksi_relaksasi = fields.Float(string='Rafaksi Relaksasi', digits=(16, 2))
    netto_relaksasi   = fields.Float(string='Tebu Final', digits=(16, 2))
    bagi_hasil        = fields.Float(string='Bagi Hasil', digits=(10, 4))
    harga_gula        = fields.Float(string='Harga Gula', digits=(16, 2))
    gula              = fields.Integer(string='Gula BH')
    rupiah            = fields.Float(string='Rupiah BH', digits=(16, 2))

    # ── Pembagian Gula BH ──
    gula_jual         = fields.Integer(string='Gula Jual (Lelang)')
    natura_kecil      = fields.Integer(string='Natura Kecil')

    # ── Natura tambahan ──
    nilai_kawalan     = fields.Float(string='Nilai Kawalan', digits=(10, 4))
    gula_kawalan      = fields.Integer(string='Natura1 (Kawalan)')
    natura2           = fields.Integer(string='Natura2 (Tetes)')
    natura3           = fields.Integer(string='Natura3 (Sharing)')

    # ── Tetes teori (disimpan, tidak ditampilkan di nota) ──
    tetes_tani        = fields.Float(string='Tetes Petani', digits=(16, 2))
    nilai_tetes       = fields.Float(string='Nilai Tetes', digits=(16, 2))

    # ── Total ──
    gula_total        = fields.Integer(string='Gula Total')
    rupiah_total      = fields.Float(string='Rupiah Total', digits=(16, 2))


class KaNtpImportRelaksasi(models.Model):
    """Staging import relaksasi per NTP."""
    _name = 'ka.ntp.import.relaksasi'
    _description = 'Import Relaksasi NTP'

    ntp_id = fields.Many2one('ka.ntp', required=True, ondelete='cascade')
    register = fields.Char(string='Register', required=True, index=True)
    persentase = fields.Float(string='Relaksasi (%)', digits=(5, 2))


class KaNtpImportBagihasil(models.Model):
    """Staging import bagi hasil per NTP."""
    _name = 'ka.ntp.import.bagihasil'
    _description = 'Import Bagi Hasil NTP'

    ntp_id = fields.Many2one('ka.ntp', required=True, ondelete='cascade')
    register = fields.Char(string='Register', required=True, index=True)
    bagi_hasil = fields.Float(string='Bagi Hasil', digits=(10, 4))


class KaNtpImportKawalan(models.Model):
    """Staging import kawalan per NTP (override nilai kawalan per register)."""
    _name = 'ka.ntp.import.kawalan'
    _description = 'Import Kawalan NTP'

    ntp_id = fields.Many2one('ka.ntp', required=True, ondelete='cascade')
    register = fields.Char(string='Register', required=True, index=True)
    kawalan = fields.Float(string='Nilai Kawalan', digits=(10, 4))


class KaNtpNotaGula(models.Model):
    """Nota Gula — agregat NTP per register."""
    _name = 'ka.ntp.nota.gula'
    _description = 'Nota Gula (Agregat per Register)'
    _order = 'register'

    ntp_id = fields.Many2one('ka.ntp', string='NTP', required=True, ondelete='cascade')
    company_id = fields.Many2one('res.company', related='ntp_id.company_id', store=True, index=True)
    register = fields.Char(string='Register', index=True)
    nama_register = fields.Char(string='Nama Register')
    petani_id = fields.Many2one('ka.petani', string='Petani')
    tebu_final = fields.Float(string='Tebu Final', digits=(16, 2))
    bagi_hasil = fields.Float(string='Bagi Hasil', digits=(10, 2))
    gula = fields.Integer(string='Gula BH')
    rupiah = fields.Float(string='Rupiah BH', digits=(16, 2))
    gula_jual = fields.Integer(string='Gula Jual (Lelang)')
    natura_kecil = fields.Integer(string='Natura Kecil')
    gula_kawalan = fields.Integer(string='Natura1 (Kawalan)')
    natura2 = fields.Integer(string='Natura2 (Tetes)')
    natura3 = fields.Integer(string='Natura3 (Sharing)')
    gula_total = fields.Integer(string='Gula Total')
    rupiah_total = fields.Float(string='Rupiah Total', digits=(16, 2))

    def action_print_kwitansi(self):
        """Cetak kwitansi untuk register ini (atau record terpilih)."""
        if not self:
            raise UserError(_('Tidak ada Nota Gula untuk dicetak.'))
        return self.env.ref('ka_sita.action_report_ka_kwitansi_ng').report_action(self)

    def action_print_kwitansi_all(self):
        """Cetak SEMUA kwitansi pada NTP yang sama dengan record ini."""
        ntps = self.mapped('ntp_id')
        all_ng = ntps.mapped('nota_gula_ids')
        if not all_ng:
            raise UserError(_('Tidak ada Nota Gula untuk dicetak.'))
        return self.env.ref('ka_sita.action_report_ka_kwitansi_ng').report_action(all_ng)
